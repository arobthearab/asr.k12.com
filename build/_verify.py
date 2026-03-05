import openpyxl
import pathlib
_BUILD = pathlib.Path(__file__).resolve().parent
wb = openpyxl.load_workbook(str(_BUILD / 'output' / 'Stride_ASR_Questionnaire_V4.xlsx'))
ws = wb['Questionnaire']

print('=== Row 2 (Classification Q) ===')
for col in 'ABCDEFGH':
    c = ws[col + '2']
    print('  %s2: %s' % (col, c.value))
print()
print('=== Row 3 ===')
for col in 'ABCDEFGH':
    c = ws[col + '3']
    print('  %s3: %s' % (col, c.value))
print()
print('=== Row 4 ===')
for col in 'ABCDEFGH':
    c = ws[col + '4']
    print('  %s4: %s' % (col, c.value))
print()
print('F2=%s' % ws['F2'].value)
print('H2=%s' % ws['H2'].value)
print('F3=%s' % ws['F3'].value)
print('F4=%s' % ws['F4'].value)
print('Max row: %d' % ws.max_row)
print('F%d=%s' % (ws.max_row, ws.cell(ws.max_row, 6).value))
print()

print('Data validations:')
for dv in ws.data_validations.dataValidation:
    f1 = dv.formula1[:80] if dv.formula1 else 'None'
    print('  range=%s  formula1=%s' % (dv.sqref, f1))
print()

ss = wb['Summary']
print('=== Summary ===')
for r in range(1, ss.max_row + 1):
    vals = []
    for c in range(1, 7):
        v = ss.cell(r, c).value
        vals.append(str(v)[:60] if v else '')
    print('  R%d: %s' % (r, vals))
