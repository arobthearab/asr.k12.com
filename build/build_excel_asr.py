#!/usr/bin/env python3
"""Stride ASR Questionnaire V4 — Excel Workbook with Live Scoring.

Generates an interactive .xlsx where assessors pick answers from drop-downs
and risk measurements are computed automatically at the question, domain,
and document levels.

Questions are loaded from asr_questions.yaml (single source of truth shared
with the Word document builder).

Scoring model
─────────────
  weight tiers       Critical = 75, High = 50, Medium = 25, Info = 10
  answer scores      Derived from risk level (L/G/M/E/H) +
                     number of non-NA choices (see score_scales in YAML)
  N/A                Always appended, score = 1, requires Notes justification
  measurement        INT(answer_score × question_weight / max_weight)
  domain aggregate   f(V,a) = ⌈Σ V_j / a^j⌉  V sorted desc, a = 4
  normalization      min(100, aggregate / 134 × 100)
"""

from pathlib import Path
from datetime import date

import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ── Load question bank from YAML ──────────────────────────────────────
YAML_PATH = Path(__file__).parent / "asr_questions.yaml"
with open(YAML_PATH, "r", encoding="utf-8") as _f:
    _QB = yaml.safe_load(_f)

WEIGHT_TIERS = _QB["weight_tiers"]
SCORE_SCALES = {int(k): v for k, v in _QB["score_scales"].items()}
NA_SCORE     = _QB["na_score"]
ASR_DOMAINS  = _QB["domains"]
CLASS_Q      = _QB["classification_question"]
WEIGHT_MAX   = max(WEIGHT_TIERS.values())  # 100

# Classification factor choices: [(text, factor), ...]
CLASS_CHOICES = [(c["text"], c["factor"]) for c in CLASS_Q["choices"]]

# ── RSK Constants ──────────────────────────────────────────────────────
RSK_DAMPING = 4
RSK_RAW_MAX = 134


# ── Style Palette ──────────────────────────────────────────────────────
_THIN  = Side(style="thin",   color="BDBDBD")
_THICK = Side(style="medium", color="757575")

TITLE_FONT    = Font(name="Calibri", bold=True, size=18, color="2E7D32")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1565C0")
HEADER_FONT   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="37474F")
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrapText=True)

DOMAIN_FONT  = Font(name="Calibri", bold=True, size=10, color="1565C0")
Q_FONT       = Font(name="Calibri", size=10)
Q_NUM_FONT   = Font(name="Calibri", bold=True, size=10)
WRAP          = Alignment(wrapText=True, vertical="top")
CENTER        = Alignment(horizontal="center", vertical="center")
NUM_FMT_0     = "0"
NUM_FMT_1     = "0.0"

DOMAIN_FILLS = [
    PatternFill("solid", fgColor="E3F2FD"),
    PatternFill("solid", fgColor="E8F5E9"),
    PatternFill("solid", fgColor="FFF3E0"),
    PatternFill("solid", fgColor="F3E5F5"),
    PatternFill("solid", fgColor="E0F7FA"),
    PatternFill("solid", fgColor="FBE9E7"),
    PatternFill("solid", fgColor="F1F8E9"),
]

WEIGHT_FILLS = {
    "Critical": PatternFill("solid", fgColor="FFCDD2"),
    "High":     PatternFill("solid", fgColor="FFE0B2"),
    "Medium":   PatternFill("solid", fgColor="C8E6C9"),
    "Info":     PatternFill("solid", fgColor="F5F5F5"),
}

ANSWER_FILL = PatternFill("solid", fgColor="FFFDE7")

GREEN_FONT  = Font(color="2E7D32", bold=True)
ORANGE_FONT = Font(color="E65100", bold=True)
RED_FONT    = Font(color="C62828", bold=True)

LOCKED   = Protection(locked=True)
UNLOCKED = Protection(locked=False)

# Gradient colours  (green → amber → red)
_GRAD_GREEN = '4CAF50'
_GRAD_AMBER = 'FFC107'
_GRAD_RED   = 'C62828'
_GRAD_MIN   = 1     # lowest meaningful measurement
_GRAD_MID   = 42    # midpoint of 1-85 range
_GRAD_MAX   = 85    # theoretical maximum measurement


def _add_risk_gradient(ws, cell_range):
    """Apply green→amber→red colour scale, keeping 0 (blank) neutral."""
    # Higher-priority rule: zero stays white (stopIfTrue prevents scale)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['0'],
                   fill=PatternFill('solid', fgColor='FFFFFF'),
                   stopIfTrue=True))
    # Three-colour gradient
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type='num', start_value=_GRAD_MIN, start_color=_GRAD_GREEN,
            mid_type='num',   mid_value=_GRAD_MID,   mid_color=_GRAD_AMBER,
            end_type='num',   end_value=_GRAD_MAX,   end_color=_GRAD_RED))


# ── Helpers ────────────────────────────────────────────────────────────

def _resolve_scores(choices):
    """Return list of (text, score) pairs including trailing N/A."""
    n = len(choices)
    scale = SCORE_SCALES[n]
    result = [(c["text"], scale[c["risk"]]) for c in choices]
    result.append(("N/A", NA_SCORE))
    return result


def _meas_formula(row, scored_choices, w_val, class_cell):
    """measurement = INT(RSK_answer/100 * RSK_weight/100 * classification_factor).

    class_cell is an absolute reference like '$H$2' holding the factor (40-100).
    Returns 0 when the answer cell is blank.
    """
    ans = "E{}".format(row)
    # Build nested-IF chain mapping answer text -> raw answer score
    expr = str(scored_choices[-1][1])
    for i in range(len(scored_choices) - 2, -1, -1):
        c_text = scored_choices[i][0].replace('"', '""')
        expr = 'IF({}="{}",{},{})'.format(ans, c_text, scored_choices[i][1], expr)
    # Three-factor product
    return '=IF({}="",0,INT({}/100*{}/100*{}))'.format(
        ans, expr, w_val, class_cell)


def _rsk_formula(rng, n_items):
    """RSK aggregate: CEILING(Sum LARGE(rng,j)/4^j, 1)."""
    terms = []
    for j in range(min(n_items, 12)):
        divisor = RSK_DAMPING ** j
        if divisor == 1:
            terms.append("LARGE({},{})".format(rng, j + 1))
        else:
            terms.append("LARGE({},{})/{}" .format(rng, j + 1, divisor))
    body = "+".join(terms)
    return "=IF(SUM({})=0,0,CEILING({},1))".format(rng, body)


def _norm_formula(rsk_cell):
    return "=IF({}=0,0,MIN(100,ROUND({}/{}*100,1)))".format(
        rsk_cell, rsk_cell, RSK_RAW_MAX)


def _norm_formula_inline(rng, n_items):
    """Residual Risk: RSK aggregate normalised to 0-100 in a single cell."""
    terms = []
    for j in range(min(n_items, 12)):
        divisor = RSK_DAMPING ** j
        if divisor == 1:
            terms.append("LARGE({},{})".format(rng, j + 1))
        else:
            terms.append("LARGE({},{})/{}" .format(rng, j + 1, divisor))
    body = "+".join(terms)
    rsk_expr = "CEILING({},1)".format(body)
    return "=IF(SUM({})=0,0,MIN(100,ROUND({}/{}*100,1)))".format(
        rng, rsk_expr, RSK_RAW_MAX)


def _rating_formula(norm_cell):
    return ('=IF({}=0,"",IF({}<=25,"Low",'
            'IF({}<=50,"Moderate",IF({}<=75,"Elevated","Critical"))))').format(
        norm_cell, norm_cell, norm_cell, norm_cell)


def _answered_formula(rng):
    return '=COUNTIF({},">0")'.format(rng)


# ── Sheet Builders ─────────────────────────────────────────────────────

def _build_instructions(wb):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "2E7D32"
    ws.column_dimensions["A"].width = 100

    r = 1
    def _w(text, font=None, gap=0):
        nonlocal r
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = Alignment(wrapText=True, vertical="top")
        cell.protection = LOCKED
        if font:
            cell.font = font
        r += 1 + gap

    _w("Stride Application Security Review \u2014 Questionnaire V4",
       TITLE_FONT, gap=1)
    _w("Interactive Excel Workbook with Live Risk Scoring",
       SUBTITLE_FONT, gap=1)

    _w("PURPOSE", Font(bold=True, size=12))
    _w("This workbook assesses the security posture of applications within "
       "the Stride technology portfolio.  Each question maps to Stride "
       "ISP/IISP policies, NIST CSF 2.0 subcategories, and, where "
       "applicable, FERPA and SOX \xa7404 ITGC requirements.")
    _w("For each question the assessor selects exactly one answer from a "
       "dropdown.  The workbook automatically computes the risk measurement "
       "per question, the aggregate per domain, and the overall "
       "questionnaire-level score.", gap=1)

    _w("HOW TO COMPLETE", Font(bold=True, size=12))
    _w("1.  Go to the Questionnaire tab.")
    _w("2.  FIRST answer the Risk Classification question at the top \u2014 "
       "this sets the global risk multiplier for ALL other measurements.",
       Font(bold=True, size=10, color="C62828"))
    _w("3.  For each domain question, click the Answer cell (yellow "
       "column E) and choose from the dropdown.")
    _w("4.  The Measurement column (F) auto-calculates immediately.")
    _w("5.  Use the Notes column (G) for justifications or context.")
    _w("6.  If you select N/A you MUST enter an explanation in the "
       "Notes column.",
       Font(bold=True, size=10, color="C62828"))
    _w("7.  When finished, review the Summary tab for domain and overall "
       "risk scores.", gap=1)

    _w("WEIGHT TIERS", Font(bold=True, size=12))
    _w("Critical ({}) \u2014 Fundamental control. Gap requires immediate "
       "CISO notification.".format(WEIGHT_TIERS["Critical"]))
    _w("High     ({}) \u2014 Important control. Gap requires 30-day "
       "remediation plan.".format(WEIGHT_TIERS["High"]))
    _w("Medium   ({}) \u2014 Supporting control. Address in next planning "
       "cycle.".format(WEIGHT_TIERS["Medium"]))
    _w("Info     ({}) \u2014 Contextual / informational only.".format(
       WEIGHT_TIERS["Info"]), gap=1)

    _w("RISK RATING SCALE", Font(bold=True, size=12))
    _w("  0\u201325%   Low       \u2014 Strong posture, controls mature "
       "and effective")
    _w(" 26\u201350%   Moderate  \u2014 Adequate posture, minor gaps with "
       "compensating controls")
    _w(" 51\u201375%   Elevated  \u2014 Material gaps requiring "
       "remediation plans")
    _w(" 76\u2013100%  Critical  \u2014 Fundamental controls missing or "
       "ineffective", gap=1)

    _w("Document Date: {}".format(date.today()),
       Font(italic=True, color="757575"))
    _w("Classification: Confidential", Font(italic=True, color="757575"))
    _w("Supersedes: ASR Questionnaire V3",
       Font(italic=True, color="757575"))

    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True


def _build_questionnaire(wb):
    """Build the Questionnaire sheet.  Returns (domain_ranges, last_data_row)."""
    ws = wb.create_sheet("Questionnaire")
    ws.sheet_properties.tabColor = "1565C0"

    # ── Column headers (row 1) ────────────────────────────────────
    headers = ["#", "Domain", "Question", "Weight", "Answer \u25bc",
               "Measurement", "Notes"]
    widths  = [4.5, 28, 58, 10, 42, 13, 32]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = Border(bottom=_THICK)
        cell.protection = LOCKED
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Transcendental classification question (row 2) ────────────
    CLASS_FILL = PatternFill("solid", fgColor="E8EAF6")   # indigo-50
    ws.cell(row=2, column=1, value="\u2605").font = Font(
        bold=True, size=12, color="C62828")
    ws.cell(row=2, column=1).alignment = CENTER
    ws.cell(row=2, column=1).protection = LOCKED
    ws.merge_cells("B2:C2")
    q_cell = ws.cell(row=2, column=2, value=CLASS_Q["text"])
    q_cell.font = Font(bold=True, size=11, color="C62828")
    q_cell.alignment = WRAP
    q_cell.protection = LOCKED
    ws.cell(row=2, column=4, value="Global").font = Font(
        bold=True, size=9, color="C62828")
    ws.cell(row=2, column=4).alignment = CENTER
    ws.cell(row=2, column=4).fill = PatternFill("solid", fgColor="FFCDD2")
    ws.cell(row=2, column=4).protection = LOCKED

    # Answer dropdown (E2) — no N/A
    class_texts = [ct[0] for ct in CLASS_CHOICES]
    ans_class = ws.cell(row=2, column=5)
    ans_class.fill = ANSWER_FILL
    ans_class.alignment = WRAP
    ans_class.protection = UNLOCKED
    dv_class = DataValidation(
        type="list",
        formula1='"' + ",".join(class_texts) + '"',
        allow_blank=True,
    )
    dv_class.error = "Please select a risk classification."
    dv_class.errorTitle = "Invalid Classification"
    dv_class.prompt = "Select the application risk classification"
    dv_class.promptTitle = "Risk Classification"
    dv_class.showErrorMessage = True
    ws.add_data_validation(dv_class)
    dv_class.add(ans_class)

    # F2: classification factor (visible, shows e.g. 40/60/80/100)
    # Nested IF mapping answer text -> factor value; 0 when blank
    factor_expr = '0'
    for txt, fval in reversed(CLASS_CHOICES):
        factor_expr = 'IF(E2="{}",{},{})'.format(
            txt.replace('"', '""'), fval, factor_expr)
    ws.cell(row=2, column=6, value='={}'.format(factor_expr))
    ws.cell(row=2, column=6).number_format = NUM_FMT_0
    ws.cell(row=2, column=6).alignment = CENTER
    ws.cell(row=2, column=6).protection = LOCKED

    # H2 (hidden col): same factor value, used as $H$2 in formulas
    ws.cell(row=2, column=8, value='=F2')
    ws.cell(row=2, column=8).protection = LOCKED
    ws.column_dimensions['H'].hidden = True

    # G2: notes (unlocked)
    ws.cell(row=2, column=7).alignment = WRAP
    ws.cell(row=2, column=7).protection = UNLOCKED

    # Tint row 2
    for c in range(1, 8):
        ws.cell(row=2, column=c).fill = CLASS_FILL
        ws.cell(row=2, column=c).border = Border(
            bottom=_THICK, left=_THIN, right=_THIN)
    ws.row_dimensions[2].height = 36

    # Reference for measurement formulas
    CLASS_CELL = '$H$2'

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A1:G1"

    # ── Domain question rows (start at row 3) ─────────────────────
    row = 3
    q_num = 0
    domain_ranges = []

    for d_idx, domain in enumerate(ASR_DOMAINS):
        start_row = row
        d_fill = DOMAIN_FILLS[d_idx % len(DOMAIN_FILLS)]

        for qi, q in enumerate(domain["questions"]):
            q_num += 1
            weight = q["weight"]
            w_val  = WEIGHT_TIERS[weight]
            scored = _resolve_scores(q["choices"])
            choice_texts = [sc[0] for sc in scored]

            # A: question number
            ws.cell(row=row, column=1, value=q_num).font = Q_NUM_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=1).protection = LOCKED

            # B: domain name (first question only)
            if qi == 0:
                ws.cell(row=row, column=2, value=domain["name"]).font = DOMAIN_FONT
                ws.cell(row=row, column=2).alignment = Alignment(
                    vertical="center", horizontal="center", wrapText=True)
            ws.cell(row=row, column=2).protection = LOCKED

            # C: question text
            ws.cell(row=row, column=3, value=q["text"]).font = Q_FONT
            ws.cell(row=row, column=3).alignment = WRAP
            ws.cell(row=row, column=3).protection = LOCKED

            # D: weight tier
            wt_cell = ws.cell(row=row, column=4, value=weight)
            wt_cell.font = Font(bold=True, size=9)
            wt_cell.fill = WEIGHT_FILLS[weight]
            wt_cell.alignment = CENTER
            wt_cell.protection = LOCKED

            # E: answer dropdown (yellow, UNLOCKED)
            ans_cell = ws.cell(row=row, column=5)
            ans_cell.fill = ANSWER_FILL
            ans_cell.alignment = WRAP
            ans_cell.protection = UNLOCKED
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(choice_texts) + '"',
                allow_blank=True,
            )
            dv.error = "Please select a valid answer from the dropdown."
            dv.errorTitle = "Invalid Answer"
            dv.prompt = "Select one of {} choices".format(len(choice_texts))
            dv.promptTitle = "Q{}".format(q_num)
            dv.showErrorMessage = True
            ws.add_data_validation(dv)
            dv.add(ans_cell)

            # F: measurement formula (locked)
            m_cell = ws.cell(row=row, column=6)
            m_cell.value = _meas_formula(row, scored, w_val, CLASS_CELL)
            m_cell.number_format = NUM_FMT_0
            m_cell.alignment = CENTER
            m_cell.protection = LOCKED

            # G: notes (UNLOCKED)
            ws.cell(row=row, column=7).alignment = WRAP
            ws.cell(row=row, column=7).protection = UNLOCKED

            # Domain tint (cols A, B, C, G)
            for c in (1, 2, 3, 7):
                ws.cell(row=row, column=c).fill = d_fill

            # Row borders
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = Border(
                    bottom=_THIN, left=_THIN, right=_THIN)

            row += 1

        end_row = row - 1
        n_q = end_row - start_row + 1

        # Merge domain name vertically
        if n_q > 1:
            ws.merge_cells(start_row=start_row, start_column=2,
                           end_row=end_row, end_column=2)

        # Thick top border for domain separator
        for c in range(1, 8):
            cell = ws.cell(row=start_row, column=c)
            existing = cell.border
            cell.border = Border(top=_THICK, bottom=existing.bottom,
                                 left=existing.left, right=existing.right)

        domain_ranges.append({
            "name":  domain["name"],
            "start": start_row,
            "end":   end_row,
            "n":     n_q,
            "pol":   ", ".join(domain.get("policy_refs", [])),
            "csf":   ", ".join(domain.get("csf_refs", [])),
        })

    last_row = row - 1

    # Conditional formatting: highlight N/A answers in orange
    for r in range(3, last_row + 1):
        ws.conditional_formatting.add(
            "E{}".format(r),
            CellIsRule(operator="equal", formula=['"N/A"'],
                       fill=PatternFill("solid", fgColor="FFE0B2")))

    # Gradient on Measurement column (F) — skip row 2 (classification factor)
    _add_risk_gradient(ws, 'F3:F{}'.format(last_row))

    # Row heights
    for r in range(3, last_row + 1):
        ws.row_dimensions[r].height = 32

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    # Sheet protection: only Answer (E) and Notes (G) are unlocked
    ws.protection.sheet = True
    ws.protection.enable()

    return domain_ranges, last_row


def _build_summary(wb, domain_ranges, last_q_row):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "C62828"

    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1,
                value="ASR Questionnaire V4 \u2014 Risk Assessment Summary")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1,
            value="Generated {}".format(date.today())).font = Font(
        italic=True, color="757575", size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=4, column=1, value="Questions Answered:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=_answered_formula(
        "Questionnaire!$F$3:$F${}".format(last_q_row)))
    ws.cell(row=4, column=3, value="of").font = Font(color="757575")
    total_q = sum(d["n"] for d in domain_ranges)
    ws.cell(row=4, column=4, value=total_q)

    ROW_H = 6
    # Columns: Domain | #Q | Answered | Residual Risk | Rating | Policy/CSF
    sum_headers = ["Domain", "# Q", "Answered",
                   "Residual Risk (0\u2013100)", "Rating",
                   "Policy / CSF References"]
    col_widths = [36, 6, 10, 18, 13, 44]
    for c, (h, w) in enumerate(zip(sum_headers, col_widths), 1):
        cell = ws.cell(row=ROW_H, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = Border(bottom=_THICK)
        ws.column_dimensions[get_column_letter(c)].width = w

    # \u2500\u2500 Domain rows (direct formulas) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    first_data_r = ROW_H + 1
    for i, dr in enumerate(domain_ranges):
        r = first_data_r + i
        d_fill = DOMAIN_FILLS[i % len(DOMAIN_FILLS)]

        ws.cell(row=r, column=1, value=dr["name"]).font = Font(
            bold=True, size=10)
        ws.cell(row=r, column=2, value=dr["n"]).alignment = CENTER

        meas_rng = "Questionnaire!$F${}:$F${}".format(
            dr["start"], dr["end"])
        ws.cell(row=r, column=3,
                value=_answered_formula(meas_rng)).alignment = CENTER

        # Residual Risk = normalised RSK (was Norm %)
        rsk_f = _rsk_formula(meas_rng, dr["n"])
        # Embed normalisation directly: MIN(100, ROUND(rsk/134*100,1))
        # We build it as a single cell so the intermediate RSK raw is hidden.
        rr_cell = ws.cell(row=r, column=4,
                          value=_norm_formula_inline(meas_rng, dr["n"]))
        rr_cell.number_format = NUM_FMT_1
        rr_cell.alignment = CENTER

        ws.cell(row=r, column=5,
                value=_rating_formula("D{}".format(r))).alignment = CENTER

        ws.cell(row=r, column=6,
                value="{}  |  {}".format(dr["pol"], dr["csf"])).font = Font(
            size=9, color="757575")

        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = d_fill
            ws.cell(row=r, column=c).border = Border(
                bottom=_THIN, left=_THIN, right=_THIN)

    # \u2500\u2500 Overall row \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    last_data_r = first_data_r + len(domain_ranges) - 1
    overall_r = last_data_r + 1
    ws.cell(row=overall_r, column=1, value="OVERALL").font = Font(
        bold=True, size=12, color="C62828")
    ws.cell(row=overall_r, column=2, value=total_q).alignment = CENTER

    overall_meas = "Questionnaire!$F$3:$F${}".format(last_q_row)
    ws.cell(row=overall_r, column=3,
            value=_answered_formula(overall_meas)).alignment = CENTER
    rr_overall = ws.cell(row=overall_r, column=4,
                         value=_norm_formula_inline(overall_meas, 12))
    rr_overall.number_format = NUM_FMT_1
    rr_overall.alignment = CENTER
    ws.cell(row=overall_r, column=5,
            value=_rating_formula(
                "D{}".format(overall_r))).alignment = CENTER

    for c in range(1, 7):
        ws.cell(row=overall_r, column=c).border = Border(
            top=_THICK, bottom=_THICK, left=_THIN, right=_THIN)
        ws.cell(row=overall_r, column=c).fill = PatternFill(
            "solid", fgColor="FFF9C4")

    # \u2500\u2500 Gradient on Residual Risk column (D) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    _add_risk_gradient(ws, 'D{}:D{}'.format(first_data_r, overall_r))

    # \u2500\u2500 Conditional formatting for rating column \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    rating_range = "E{}:E{}".format(first_data_r, overall_r)
    ws.conditional_formatting.add(rating_range, CellIsRule(
        operator="equal", formula=['"Low"'],
        font=GREEN_FONT,
        fill=PatternFill("solid", fgColor="E8F5E9")))
    ws.conditional_formatting.add(rating_range, CellIsRule(
        operator="equal", formula=['"Moderate"'],
        font=Font(bold=True),
        fill=PatternFill("solid", fgColor="FFF3E0")))
    ws.conditional_formatting.add(rating_range, CellIsRule(
        operator="equal", formula=['"Elevated"'],
        font=ORANGE_FONT,
        fill=PatternFill("solid", fgColor="FFE0B2")))
    ws.conditional_formatting.add(rating_range, CellIsRule(
        operator="equal", formula=['"Critical"'],
        font=RED_FONT,
        fill=PatternFill("solid", fgColor="FFCDD2")))

    # Rating scale reference
    ref_r = overall_r + 2
    ws.cell(row=ref_r, column=1,
            value="Risk Rating Scale").font = Font(bold=True, size=11)
    scale_data = [
        ("0\u201325%",   "Low",      "Strong posture \u2014 controls mature and effective",   "E8F5E9"),
        ("26\u201350%",  "Moderate",  "Adequate \u2014 minor gaps with compensating controls", "FFF3E0"),
        ("51\u201375%",  "Elevated",  "Material gaps requiring remediation plans",              "FFE0B2"),
        ("76\u2013100%", "Critical",  "Fundamental controls missing or ineffective",            "FFCDD2"),
    ]
    for j, (pct, lbl, desc, clr) in enumerate(scale_data):
        rr = ref_r + 1 + j
        ws.cell(row=rr, column=1, value=pct).alignment = CENTER
        ws.cell(row=rr, column=2, value=lbl).font = Font(bold=True)
        ws.cell(row=rr, column=3, value=desc)
        for c in range(1, 4):
            ws.cell(row=rr, column=c).fill = PatternFill(
                "solid", fgColor=clr)

    # App info block
    info_r = ref_r + 6
    ws.cell(row=info_r, column=1,
            value="Application Information").font = Font(
        bold=True, size=11)
    fields = ["Application Name", "Business Owner", "Technical Owner",
              "CMDB ID", "Risk Classification", "Assessment Date",
              "Assessor"]
    for j, fld in enumerate(fields):
        rr = info_r + 1 + j
        ws.cell(row=rr, column=1,
                value=fld + ":").font = Font(bold=True, size=10)
        entry = ws.cell(row=rr, column=2)
        entry.fill = ANSWER_FILL
        entry.border = Border(
            bottom=Side(style="thin", color="757575"))
        entry.protection = UNLOCKED
        ws.merge_cells(start_row=rr, start_column=2,
                       end_row=rr, end_column=4)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.enable()


def _build_scoring_model(wb):
    """Reference sheet with scoring mechanics."""
    ws = wb.create_sheet("Scoring Model")
    ws.sheet_properties.tabColor = "757575"
    ws.column_dimensions["A"].width = 90

    r = 1
    def _w(text, font=None, gap=0):
        nonlocal r
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = Alignment(wrapText=True, vertical="top")
        cell.protection = LOCKED
        if font:
            cell.font = font
        r += 1 + gap

    _w("ASR Scoring Model \u2014 Reference", TITLE_FONT, gap=1)

    _w("RISK CLASSIFICATION (TRANSCENDENTAL QUESTION)", Font(bold=True, size=12))
    _w("The first question on the Questionnaire asks for the application's "
       "risk classification.  The answer sets a global multiplier that "
       "scales EVERY other question's measurement:")
    for txt, fval in CLASS_CHOICES:
        _w("  {:30s}  factor = {}".format(txt, fval),
           Font(name="Consolas", size=10))
    _w("N/A is NOT available for this question \u2014 risk classification "
       "is always applicable.", gap=1)

    _w("WEIGHT TIERS", Font(bold=True, size=12))
    _w("Each question is assigned a weight tier reflecting its importance:")
    for tier, val in WEIGHT_TIERS.items():
        _w("  {:10s}  weight = {}".format(tier, val),
           Font(name="Consolas", size=10))
    _w("")

    _w("ANSWER SCORING", Font(bold=True, size=12))
    _w("Each non-N/A answer carries a risk level (L, G, M, E, H). "
       "The numeric score depends on the number of non-N/A choices:")
    _w("")
    for n_choices in sorted(SCORE_SCALES):
        sc = SCORE_SCALES[n_choices]
        labels = ", ".join("{}={}".format(k, v) for k, v in sc.items())
        _w("  {} choices: {}".format(n_choices, labels),
           Font(name="Consolas", size=10))
    _w("  N/A = {}  (always available; requires justification "
       "in Notes)".format(NA_SCORE),
       Font(name="Consolas", size=10))
    _w("")

    _w("PER-QUESTION MEASUREMENT", Font(bold=True, size=12))
    _w("  measurement = INT( RSK(answer)/100  x  RSK(weight)/100  "
       "x  RSK(classification) )",
       Font(name="Consolas", size=10))
    _w("Three normalised factors multiplied:")
    _w("  \u2022 RSK(answer)         = answer risk score (1\u201385)")
    _w("  \u2022 RSK(weight)         = question weight tier (13\u2013100)")
    _w("  \u2022 RSK(classification) = classification factor (40\u2013100)")
    _w("Theoretical max = INT(85/100 x 100/100 x 100) = 85", gap=1)

    _w("AGGREGATE SCORING", Font(bold=True, size=12))
    _w("Domain and overall aggregates use a diminishing-impact "
       "composite where the most severe finding dominates:")
    _w("  Composite = CEILING( Sum meas_sorted_desc / damping^j , 1)",
       Font(name="Consolas", size=10))
    _w("This prevents inflation from many minor gaps.", gap=1)

    _w("NORMALIZATION", Font(bold=True, size=12))
    _w("  norm_pct = min(100, raw_aggregate / {} x 100)".format(
       RSK_RAW_MAX), Font(name="Consolas", size=10))
    _w("")

    _w("RISK RATING THRESHOLDS", Font(bold=True, size=12))
    _w("  0\u201325%   Low       \u2014 Strong posture")
    _w(" 26\u201350%   Moderate  \u2014 Adequate, minor gaps")
    _w(" 51\u201375%   Elevated  \u2014 Material gaps")
    _w(" 76\u2013100%  Critical  \u2014 Fundamental gaps")

    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True


# ── Main ───────────────────────────────────────────────────────────────

def build_workbook():
    wb = Workbook()

    _build_instructions(wb)
    domain_ranges, last_row = _build_questionnaire(wb)
    _build_summary(wb, domain_ranges, last_row)
    _build_scoring_model(wb)

    wb.active = wb.sheetnames.index("Questionnaire")

    out = Path(__file__).resolve().parent / "output" / "Stride_ASR_Questionnaire_V4.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    sz = out.stat().st_size
    print("Saved: {} ({:,} bytes)".format(out, sz))
    return out


if __name__ == "__main__":
    build_workbook()
