#!/usr/bin/env python3
"""One-shot generator: writes the new build_excel_asr.py content."""
import pathlib, textwrap

TARGET = pathlib.Path(__file__).resolve().parent / "build_excel_asr.py"

CONTENT = textwrap.dedent('''\
#!/usr/bin/env python3
"""Stride ASR Questionnaire V4 \u2014 Excel Workbook with Live Scoring.

Generates an interactive .xlsx where assessors pick answers from drop-downs
and risk measurements are computed automatically at the question, domain,
and document levels.

Questions are loaded from asr_questions.yaml (single source of truth shared
with the Word document builder).

Scoring model
\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
  weight tiers       Critical = 75, High = 50, Medium = 25, Info = 10
  answer scores      Derived from risk level (L/G/M/E/H) +
                     number of non-NA choices (see score_scales in YAML)
  N/A                Always appended, score = 1, requires Notes justification
  measurement        INT(answer_score \u00d7 question_weight / max_weight)
  domain aggregate   f(V,a) = \u2308\u03a3 V_j / a^j\u2309  V sorted desc, a = 4
  normalization      min(100, aggregate / 134 \u00d7 100)
"""

import math
from pathlib import Path
from datetime import date

import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# \u2500\u2500 Load question bank from YAML \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
YAML_PATH = Path(__file__).parent / "asr_questions.yaml"
with open(YAML_PATH, "r", encoding="utf-8") as _f:
    _QB = yaml.safe_load(_f)

WEIGHT_TIERS = _QB["weight_tiers"]
SCORE_SCALES = {int(k): v for k, v in _QB["score_scales"].items()}
NA_SCORE     = _QB["na_score"]
ASR_DOMAINS  = _QB["domains"]
WEIGHT_MAX   = max(WEIGHT_TIERS.values())  # 75

# RSK Constants
RSK_DAMPING = 4
RSK_RAW_MAX = 134


# \u2500\u2500 Style Palette \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
_THIN  = Side(style="thin",   color="BDBDBD")
_THICK = Side(style="medium", color="757575")

TITLE_FONT    = Font(name="Calibri", bold=True, size=18, color="2E7D32")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1565C0")
HEADER_FONT   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="37474F")
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrapText=True)

DOMAIN_FONT  = Font(name="Calibri", bold=True, size=10, color="1565C0")
Q_FONT       = Font(name="Calibri", size=10)
Q_NUM_FONT   = Font(name="Calibri", bold=True, size=10)
WRAP          = Alignment(wrapText=True, vertical="top")
CENTER        = Alignment(horizontal="center", vertical="center")
NUM_FMT_0     = "0"
NUM_FMT_1     = "0.0"

DOMAIN_FILLS = [
    PatternFill("solid", fgColor="E3F2FD"),
    PatternFill("solid", fgColor="E8F5E9"),
    PatternFill("solid", fgColor="FFF3E0"),
    PatternFill("solid", fgColor="F3E5F5"),
    PatternFill("solid", fgColor="E0F7FA"),
    PatternFill("solid", fgColor="FBE9E7"),
    PatternFill("solid", fgColor="F1F8E9"),
]

WEIGHT_FILLS = {
    "Critical": PatternFill("solid", fgColor="FFCDD2"),
    "High":     PatternFill("solid", fgColor="FFE0B2"),
    "Medium":   PatternFill("solid", fgColor="C8E6C9"),
    "Info":     PatternFill("solid", fgColor="F5F5F5"),
}

ANSWER_FILL = PatternFill("solid", fgColor="FFFDE7")

GREEN_FONT  = Font(color="2E7D32", bold=True)
ORANGE_FONT = Font(color="E65100", bold=True)
RED_FONT    = Font(color="C62828", bold=True)

LOCKED   = Protection(locked=True)
UNLOCKED = Protection(locked=False)


# \u2500\u2500 Helpers \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500

def _resolve_scores(choices):
    """Return list of (text, score) pairs including trailing N/A."""
    n = len(choices)
    scale = SCORE_SCALES[n]
    result = [(c["text"], scale[c["risk"]]) for c in choices]
    result.append(("N/A", NA_SCORE))
    return result


def _meas_formula(row, scored_choices, w_val):
    """Nested-IF: answer -> INT(score * weight / max_weight). 0 when blank."""
    ans = f"E{row}"
    expr = str(scored_choices[-1][1])
    for i in range(len(scored_choices) - 2, -1, -1):
        c_text = scored_choices[i][0].replace(\\'"\\', \\'""\\')
        expr = f\\'IF({ans}="{c_text}",{scored_choices[i][1]},{expr})\\'
    return f\\'=IF({ans}="",0,INT({expr}*{w_val}/{WEIGHT_MAX}))\\'


def _rsk_formula(rng, n_items):
    """RSK aggregate: CEILING(Sum LARGE(rng,j)/4^j, 1)."""
    terms = []
    for j in range(min(n_items, 12)):
        divisor = RSK_DAMPING ** j
        if divisor == 1:
            terms.append(f"LARGE({rng},{j+1})")
        else:
            terms.append(f"LARGE({rng},{j+1})/{divisor}")
    body = "+".join(terms)
    return f\\'=IF(SUM({rng})=0,0,CEILING({body},1))\\'


def _norm_formula(rsk_cell):
    return f\\'=IF({rsk_cell}=0,0,MIN(100,ROUND({rsk_cell}/{RSK_RAW_MAX}*100,1)))\\'


def _rating_formula(norm_cell):
    return (f\\'=IF({norm_cell}=0,"",IF({norm_cell}<=25,"Low",\\'
            f\\'IF({norm_cell}<=50,"Moderate",IF({norm_cell}<=75,"Elevated","Critical"))))\\'
            )


def _answered_formula(rng):
    return f\\'=COUNTIF({rng},">0")\\'
''')

print("ERROR: triple-quote nesting prevents this approach too.")
print("Use direct file editing instead.")
